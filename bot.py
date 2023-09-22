import sys
import random
import openpyxl

#from joblib import dump, load
from sklearn.model_selection import train_test_split  
from sklearn.metrics import accuracy_score, f1_score
from sklearn.feature_extraction.text import CountVectorizer
from sklearn.ensemble import GradientBoostingClassifier


INTENTS = {}

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("data.xlsx")
# Define variable to read the active sheet:
worksheet = wookbook.active
# Iterate the loop to read the cell values
for row in range(2, worksheet.max_row):
    intent = worksheet.cell(row=row, column=2).value.lower()
    question = worksheet.cell(row=row, column=3).value.lower()
    answer = worksheet.cell(row=row, column=4).value.lower()
    if (not INTENTS.get(intent)):
        INTENTS[intent] = []
    
    INTENTS[intent].append({"question": question, "answer": answer})


XIntension = []
YIntension = []

XAnswer = []
YAnswer = []

for intent in INTENTS:
    questions = [item['question'] for item in INTENTS[intent]]
    for question in questions:
        XIntension.append(question)
        YIntension.append(intent)

for intent in INTENTS:
    for data in INTENTS[intent]:
        XAnswer.append(data['question'])
        YAnswer.append(data['answer'])


vectorizer = CountVectorizer()
vectorizer.fit(XIntension)
#dump(vectorizer, 'vector.joblib')
vecX = vectorizer.transform(XIntension)

vectorizer.fit(XAnswer)
vecXanswer = vectorizer.transform(XAnswer)

modelIntent = GradientBoostingClassifier()
modelIntent.fit(vecX, YIntension)

modelAnswer = GradientBoostingClassifier()
modelAnswer.fit(vecXanswer, YAnswer)

#dump(model, "intents.joblib")


#model = load("intents.joblib")
#vectorizer = load("vector.joblib")

if __name__ == "__main__":
    print(sys.argv[1])
    if len (sys.argv) > 1:
        predictions = modelIntent.predict(vectorizer.transform([sys.argv[1].lower().replace(" ", "")]))
        intent = predictions[0]

        print(intent)

        predictions = modelAnswer.predict(vectorizer.transform([sys.argv[1].lower().replace(" ", "")]))

        print(predictions[0])

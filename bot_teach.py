import openpyxl
import yake
import json

import re

from joblib import dump
from sklearn.model_selection import train_test_split
from sklearn.feature_extraction.text import TfidfVectorizer, CountVectorizer
from sklearn.ensemble import GradientBoostingClassifier
from sklearn.svm import LinearSVC


INTENTS = {}

QUESTION_EXCLUDE = [
    'что',
    'как',
    'почему',
    'такое',
    'где',
    'можно',
    'найти',
    'делать',
    'возможно',
    'можно',
    'возможно-ли',
    'можно-ли',
    'в',
    'каких',
    'какие',
    'кто',
    'если',
    'существуют',
    'для',
    'портал',
    'портале',
    'поставщик',
    'поставщиков',
    'при',
    'через',
]

QUESTION_NOT_EXCLUDE = [
    '44',
    '223',
    '44 фз',
    '223 фз',
    'фз 44',
    'фз 223',
    '44-фз',
    '223-фз',
    'фз-44',
    'фз-223',
    'кс',
    'окпд2',
    '44-ФЗ',
    '94-ФЗ',
    'АРМ',
    'ЕИС',
    'ЕАИСТ',
    'ИНН',
    'ИСиР',
    'КПГЗ',
    'КПП',
    'КТС',
    'НДС',
    'НМЦК',
    'ОГРН',
    'ОКПД',
    'ОКТМО',
    'ОПО',
    'СТЕ',
    'СУДИР',
    'УПД',
    'ЭМ',
    'МО',
    'ЕАСУЗ'
]

ADDITIONAL_QUESTIONS = {
    "упд": "УПД — это универсальный передаточный документ. Его особенность — многофункциональность, благодаря которой можно заметно уменьшить объем документооборота."
}



extractor = yake.KeywordExtractor(lan='ru', stopwords=QUESTION_EXCLUDE, n=3)


# as per recommendation from @freylis, compile once only
CLEANR = re.compile('<.*?>') 

def cleanhtml(raw_html):
  cleantext = re.sub(CLEANR, '', raw_html)
  return cleantext


def fillDatasheet(intent, question, answer):
    if (intent):
        intent = intent.lower()
        question = question.lower()
        answer = answer.lower()

        keywords = extractor.extract_keywords(question)
        parts = question.lower().split(" ")
        notExcludes = [part for part in parts if part in [x.lower() for x in QUESTION_NOT_EXCLUDE]]

        for keyword in keywords:
            if (not INTENTS.get(intent)):
                INTENTS[intent] = []
            INTENTS[intent].append({"question": str(keyword[0]), "answer": answer})

        for notExclude in notExcludes:
            if (not INTENTS.get(intent)):
                INTENTS[intent] = []
            INTENTS[intent].append({"question": str(notExclude), "answer": answer})
        

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("data.xlsx")
# Define variable to read the active sheet:
worksheet = wookbook.active
# Iterate the loop to read the cell values
for row in range(2, worksheet.max_row):
    intent = worksheet.cell(row=row, column=2).value
    question = worksheet.cell(row=row, column=3).value
    answer = worksheet.cell(row=row, column=4).value
    fillDatasheet(intent, question, answer)

# Define variable to load the wookbook
wookbook = openpyxl.load_workbook("data2.xlsx")
# Define variable to read the active sheet:
worksheet = wookbook.active
# Iterate the loop to read the cell values
for row in range(2, worksheet.max_row):
    intent = worksheet.cell(row=row, column=3).value
    question = worksheet.cell(row=row, column=2).value
    answer = worksheet.cell(row=row, column=5).value
    fillDatasheet(intent, question, answer)    

for additional in ADDITIONAL_QUESTIONS:
    fillDatasheet("доплнения", additional.lower(), ADDITIONAL_QUESTIONS[additional].lower()) 


# Opening JSON file
f = open('data3.json', encoding='utf-8')
 
# returns JSON object as
# a dictionary
data = json.load(f)

for row in data:
    for service in row['servicesList']:
        for article in service['articleList']:
            intent = row['categoryName']
            question = article['nameLarge']
            answer = cleanhtml(article['previewText'])
            fillDatasheet(intent, question, answer)    

# Opening JSON file
f = open('data4.json', encoding='utf-8')
 
# returns JSON object as
# a dictionary
data = json.load(f)

for row in data:
    intent = 'Термины и сокращения'
    question = row
    answer = data[row]
    fillDatasheet(intent, question, answer)  

XAnswer = []
YAnswer = []

for intent in INTENTS:
    for data in INTENTS[intent]:
        XAnswer.append(data['question'])
        YAnswer.append(data['answer'])

vectorizer = GradientBoostingClassifier(analyzer='char_wb', ngram_range=(2,3), max_df=0.8)
vectorizer.fit(XAnswer)
dump(vectorizer, 'vector-answers.joblib')

vecXanswer = vectorizer.transform(XAnswer)

modelAnswer = LinearSVC()
modelAnswer.fit(vecXanswer, YAnswer)

dump(modelAnswer, "answers.joblib")


